from pathlib import Path
import pandas as pd

BASE = Path(".")
OUT = BASE / "reports" / "avances_dataset_normalizado_redatam.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

def read_csv_safe(path):
    path = Path(path)
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()

entities = read_csv_safe("data/checks/redatamx/redatam_entities_all.csv")
vars_all = read_csv_safe("data/checks/redatamx/redatam_variables_all.csv")
escolaridad = read_csv_safe("data/checks/redatamx/selection/escolaridad_candidatas.csv")
vivienda = read_csv_safe("data/checks/redatamx/selection/vivienda_candidatas.csv")
dimensiones = read_csv_safe("data/checks/redatamx/selection/dimensiones_candidatas.csv")
summary_entity = read_csv_safe("data/checks/redatamx/summary_variables_by_entity.csv")

resumen = pd.DataFrame([
    {
        "campo": "Objetivo del avance",
        "valor": "Definir dataset normalizado para proyecto BI con Redatam"
    },
    {
        "campo": "Países seleccionados",
        "valor": "Chile, Brasil, Argentina, Perú"
    },
    {
        "campo": "Temas seleccionados",
        "valor": "Escolaridad y Vivienda"
    },
    {
        "campo": "Herramienta usada",
        "valor": "redatamx + DuckDB + Python"
    },
    {
        "campo": "Estado",
        "valor": "Variables reales extraídas desde diccionarios .dicx; falta validar códigos/categorías"
    },
    {
        "campo": "Regla metodológica",
        "valor": "No se comparan nombres originales; se homologan conceptos e indicadores"
    },
    {
        "campo": "Dataset final esperado",
        "valor": "Tabla agregada por país, geografía, sexo, edad_tramo, tema, indicador y categoría"
    },
])

paises = pd.DataFrame([
    {
        "country_id": "CHL",
        "country_name": "Chile",
        "dataset": "CP2017CHL",
        "census_year": 2017,
        "persona_entity": "PERSONA",
        "hogar_entity": "HOGAR",
        "vivienda_entity": "VIVIENDA",
        "geo_level_recomendado": "COMUNA",
        "estado": "diccionario abierto con redatamx"
    },
    {
        "country_id": "BRA",
        "country_name": "Brasil",
        "dataset": "CP2010BRA",
        "census_year": 2010,
        "persona_entity": "PESSOA",
        "hogar_entity": "FAMILIA",
        "vivienda_entity": "DOMICIL",
        "geo_level_recomendado": "MUNIC",
        "estado": "diccionario abierto con redatamx"
    },
    {
        "country_id": "ARG",
        "country_name": "Argentina",
        "dataset": "CP2010ARG",
        "census_year": 2010,
        "persona_entity": "PERSONA",
        "hogar_entity": "HOGAR",
        "vivienda_entity": "VIVIENDA",
        "geo_level_recomendado": "DPTO",
        "estado": "diccionario abierto con redatamx"
    },
    {
        "country_id": "PER",
        "country_name": "Perú",
        "dataset": "CP2017PER",
        "census_year": 2017,
        "persona_entity": "PERSONA",
        "hogar_entity": "HOGAR",
        "vivienda_entity": "VIVIENDA",
        "geo_level_recomendado": "DISTRITO",
        "estado": "diccionario abierto con redatamx"
    },
])

formato = pd.DataFrame([
    {
        "campo_normalizado": "country_id",
        "descripcion": "Código de país",
        "tipo": "texto",
        "valores": "CHL, BRA, ARG, PER",
        "fuente": "definido por el grupo"
    },
    {
        "campo_normalizado": "country_name",
        "descripcion": "Nombre de país",
        "tipo": "texto",
        "valores": "Chile, Brasil, Argentina, Perú",
        "fuente": "definido por el grupo"
    },
    {
        "campo_normalizado": "census_year",
        "descripcion": "Año censal del dataset",
        "tipo": "entero",
        "valores": "2010 o 2017 según país",
        "fuente": "dataset Redatam"
    },
    {
        "campo_normalizado": "geo_id",
        "descripcion": "Código territorial",
        "tipo": "texto",
        "valores": "según país",
        "fuente": "entidad geográfica Redatam"
    },
    {
        "campo_normalizado": "geo_name",
        "descripcion": "Nombre territorial",
        "tipo": "texto",
        "valores": "comuna/municipio/departamento/distrito",
        "fuente": "entidad geográfica Redatam"
    },
    {
        "campo_normalizado": "geo_level",
        "descripcion": "Nivel territorial",
        "tipo": "texto",
        "valores": "comuna, municipio, departamento, distrito",
        "fuente": "homologación"
    },
    {
        "campo_normalizado": "sexo_homologado",
        "descripcion": "Sexo de la persona",
        "tipo": "categoría",
        "valores": "hombre, mujer, no_informado",
        "fuente": "PERSONA/PESSOA"
    },
    {
        "campo_normalizado": "edad_tramo",
        "descripcion": "Edad agrupada",
        "tipo": "categoría",
        "valores": "0_14, 15_24, 25_44, 45_64, 65_mas, no_informado",
        "fuente": "PERSONA/PESSOA"
    },
    {
        "campo_normalizado": "tamano_poblacion_segmento",
        "descripcion": "Segmento de tamaño poblacional del territorio",
        "tipo": "categoría",
        "valores": "muy_pequeno, pequeno, mediano, grande, metropolitano",
        "fuente": "calculado por agregación territorial"
    },
    {
        "campo_normalizado": "tema",
        "descripcion": "Tema del análisis",
        "tipo": "categoría",
        "valores": "escolaridad, vivienda",
        "fuente": "definido por el grupo"
    },
    {
        "campo_normalizado": "indicador",
        "descripcion": "Indicador BI calculado",
        "tipo": "texto",
        "valores": "secundaria_completa_o_mas, superior_completa_o_mas, agua_adecuada, saneamiento_adecuado, hacinamiento",
        "fuente": "homologación"
    },
    {
        "campo_normalizado": "category",
        "descripcion": "Categoría del indicador",
        "tipo": "texto",
        "valores": "si, no, no_informado u otra categoría homologada",
        "fuente": "recodificación"
    },
    {
        "campo_normalizado": "numerator",
        "descripcion": "Casos que cumplen condición",
        "tipo": "numérico",
        "valores": "conteo o conteo ponderado",
        "fuente": "consulta Redatam"
    },
    {
        "campo_normalizado": "denominator",
        "descripcion": "Población base del indicador",
        "tipo": "numérico",
        "valores": "conteo o conteo ponderado",
        "fuente": "consulta Redatam"
    },
    {
        "campo_normalizado": "pct",
        "descripcion": "Porcentaje",
        "tipo": "decimal",
        "valores": "numerator / denominator",
        "fuente": "calculado"
    },
    {
        "campo_normalizado": "weight_used",
        "descripcion": "Indica si se usó ponderador",
        "tipo": "booleano",
        "valores": "true, false",
        "fuente": "según país/entidad"
    },
    {
        "campo_normalizado": "source_variable",
        "descripcion": "Variable original Redatam usada",
        "tipo": "texto",
        "valores": "nombre original por país",
        "fuente": "redatam_variables_all.csv"
    },
    {
        "campo_normalizado": "quality_flag",
        "descripcion": "Estado metodológico",
        "tipo": "categoría",
        "valores": "ok, revisar_categoria, no_comparable, pendiente",
        "fuente": "validación del grupo"
    },
])

indicadores = pd.DataFrame([
    {
        "tema": "escolaridad",
        "indicador": "secundaria_completa_o_mas",
        "descripcion": "% de población de 25 años o más con secundaria/media completa o nivel superior",
        "base": "personas de 25 años o más",
        "estado": "pendiente validar categorías"
    },
    {
        "tema": "escolaridad",
        "indicador": "superior_completa_o_mas",
        "descripcion": "% de población de 25 años o más con educación superior completa o más",
        "base": "personas de 25 años o más",
        "estado": "pendiente validar categorías"
    },
    {
        "tema": "escolaridad",
        "indicador": "baja_escolaridad",
        "descripcion": "% de población de 15 años o más sin escolaridad o con primaria/básica incompleta",
        "base": "personas de 15 años o más",
        "estado": "pendiente validar categorías"
    },
    {
        "tema": "vivienda",
        "indicador": "agua_adecuada",
        "descripcion": "% de hogares/viviendas con acceso adecuado a agua",
        "base": "hogares o viviendas particulares",
        "estado": "pendiente validar categorías"
    },
    {
        "tema": "vivienda",
        "indicador": "saneamiento_adecuado",
        "descripcion": "% de hogares/viviendas con baño, desagüe, cloaca o alcantarillado adecuado",
        "base": "hogares o viviendas particulares",
        "estado": "pendiente validar categorías"
    },
    {
        "tema": "vivienda",
        "indicador": "hacinamiento",
        "descripcion": "% de hogares con hacinamiento según variable directa o personas por dormitorio",
        "base": "hogares",
        "estado": "pendiente validar categorías"
    },
])

pendientes = pd.DataFrame([
    {
        "prioridad": "alta",
        "pendiente": "Revisar códigos/categorías de variables seleccionadas",
        "motivo": "Sin categorías no se pueden calcular indicadores comparables"
    },
    {
        "prioridad": "alta",
        "pendiente": "Confirmar variable de escolaridad principal por país",
        "motivo": "Debe mapearse a nivel_educativo_homologado"
    },
    {
        "prioridad": "alta",
        "pendiente": "Confirmar variables de agua, saneamiento y hacinamiento por país",
        "motivo": "Serán indicadores principales de vivienda"
    },
    {
        "prioridad": "media",
        "pendiente": "Confirmar ponderadores por país",
        "motivo": "Brasil y Argentina pueden requerir ponderación"
    },
    {
        "prioridad": "media",
        "pendiente": "Construir tamaño poblacional segmentado",
        "motivo": "Dimensión mínima solicitada por el trabajo"
    },
    {
        "prioridad": "media",
        "pendiente": "Probar consulta agregada con redatamx en Chile",
        "motivo": "Chile será base para GeoDa"
    },
])

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    resumen.to_excel(writer, sheet_name="01_Resumen", index=False)
    paises.to_excel(writer, sheet_name="02_Paises_Datasets", index=False)
    entities.to_excel(writer, sheet_name="03_Entidades", index=False)
    escolaridad.to_excel(writer, sheet_name="04_Variables_Escolaridad", index=False)
    vivienda.to_excel(writer, sheet_name="05_Variables_Vivienda", index=False)
    dimensiones.to_excel(writer, sheet_name="06_Dimensiones", index=False)
    formato.to_excel(writer, sheet_name="07_Formato_Normalizado", index=False)
    indicadores.to_excel(writer, sheet_name="08_Indicadores", index=False)
    pendientes.to_excel(writer, sheet_name="09_Pendientes", index=False)
    summary_entity.to_excel(writer, sheet_name="10_Resumen_Entidades", index=False)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook(OUT)

header_fill = PatternFill("solid", fgColor="D9EAF7")
title_fill = PatternFill("solid", fgColor="EEF5FB")
thin = Side(style="thin", color="D9E2EC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row >= 2 and max_col >= 1:
        table_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        safe_name = "T_" + "".join(ch for ch in ws.title if ch.isalnum())
        table = Table(displayName=safe_name[:250], ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col[:200]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

    ws.row_dimensions[1].height = 28

wb.save(OUT)
print(f"Excel creado: {OUT}")
